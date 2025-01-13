from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, colors, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import Counter
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor


class CreateExcelTools:

    def __init__(self, excel_name, wrapText=False):
        """
        初始化
        :param excel_name: excel路径
        :param wrapText: 是否自动换行
        """
        self.sheet_dict = {}
        self.sheet_row_index = {}
        self.excel_obj = Workbook()
        self.excel_name = excel_name
        self.sheet_index = 0
        # 设置居中所有单元格
        self.align = Alignment(horizontal='center', vertical='center', wrapText=wrapText)
        # 设置边框
        self.border = Border(
            top=Side(border_style='medium', color=colors.BLACK),
            bottom=Side(border_style='medium', color=colors.BLACK),
            left=Side(border_style='medium', color=colors.BLACK),
            right=Side(border_style='medium', color=colors.BLACK)
        )

    def create_sheet(self, sheet_name):
        """创建sheet"""
        if sheet_name not in self.sheet_dict.keys():
            self.sheet_dict[sheet_name] = self.excel_obj.create_sheet(sheet_name, self.sheet_index)
            self.sheet_index += 1
        self.sheet_row_index[sheet_name] = 1

    def sheet(self, sheet_name):
        return self.sheet_dict[sheet_name]

    @classmethod
    def save_sample_data(cls, excel_name, all_data_dict):
        """
        存储列表数据(纯文本)
        :param all_data_dict: {
            'sheetname1': [["name", "age", "salary"],
                           ["dd", 14, "2"]],
            'sheetname2': [["name", "age", "salary"],
                           ["dd", 15, "2"]]
        }
        """
        excel_obj = Workbook()
        index = 0
        for sheet_name, all_data in all_data_dict.items():
            ws = excel_obj.create_sheet(sheet_name, index)
            for data in all_data:
                ws.append(data)
            index += 1
        excel_obj.save(excel_name)

    def set_freeze_panes(self, sheet_name, freeze_panes=None):
        """
        设置冻结窗格
        :param sheet_name: sheet名
        :param freeze_panes: 冻结位置信息，如：'A2'
        :return:
        """
        self.sheet(sheet_name).freeze_panes = self.sheet(sheet_name)[freeze_panes]

    def insert_image(self, current_sheet, last_img_width, img_path, current_row_index, column_index):
        """
        将图片插入到指定单元格，并使用 TwoCellAnchor 进行锚定。

        :param current_sheet: 要插入图片的工作表
        :param last_img_width: 上一次插入图片的累积宽度
        :param img_path: 图片文件的路径
        :param current_row_index: 单元格横坐标
        :param column_index: 单元格纵坐标
        """

        # 创建图片对象
        img = Image(img_path)

        img_width, img_height = img.width, img.height
        cell_width = current_sheet.column_dimensions[get_column_letter(column_index)].width
        cell_height = current_sheet.row_dimensions[current_row_index].height
        target_image_width = (img_width / 7.2) + 10
        if cell_width:
            if cell_width < target_image_width:
                current_sheet.column_dimensions[
                    get_column_letter(column_index)].width = target_image_width + last_img_width
        else:
            current_sheet.column_dimensions[get_column_letter(column_index)].width = target_image_width + last_img_width
        target_image_height = img_height / 0.98
        if cell_height:

            if cell_height < target_image_height:
                current_sheet.row_dimensions[current_row_index].height = target_image_height
        else:
            current_sheet.row_dimensions[current_row_index].height = target_image_height

        # 计算起始和结束锚点
        # 起始单元格锚点
        start_marker = AnchorMarker(col=column_index - 1, colOff=9525 * (10 + last_img_width),
                                    row=current_row_index - 1,
                                    rowOff=9525 * 10)
        # 结束单元格锚点
        end_marker = AnchorMarker(col=column_index - 1, colOff=9525 * (img_width + last_img_width),
                                  row=current_row_index - 1, rowOff=9525 * img_height)

        # 创建 TwoCellAnchor 对象
        img.anchor = TwoCellAnchor("twoCell", start_marker, end_marker)
        # 将 TwoCellAnchor 添加到 worksheet 中
        current_sheet.add_image(img)
        return img_width + last_img_width

    def add_row_data(self, sheet_name, row_datas, is_center=False, add_border=False, width_default=None,
                     height_default=None):
        """
        添加行数据
        :param row_datas:行数据，需填写文本值，可选width,height,color,font,合并单元格信息
        :param is_center:是否设置文本居中
        :param add_border:是否添加边框
        :param width_default:单元格默认宽度
        :param height_default:单元格默认高度
        :return:
        """
        current_sheet = self.sheet(sheet_name)
        current_row_index = self.sheet_row_index[sheet_name]
        for column_index, row_data in enumerate(row_datas, start=1):
            cell_coordinate = f"{get_column_letter(column_index)}{current_row_index}"
            current_cell = current_sheet[cell_coordinate]
            if current_cell.coordinate in current_sheet.merged_cells:
                continue
            current_cell.value = row_data.get('value')  # 设置值
            if is_center:
                current_cell.alignment = self.align  # 设置单元格内容剧中
            if add_border:
                current_cell.border = self.border
            # 设置单元格高度
            height = row_data.get("height")
            if height:
                current_sheet.row_dimensions[current_row_index].height = height
            elif height_default:
                current_sheet.row_dimensions[current_row_index].height = height_default
            # 设置单元格宽度
            width = row_data.get("width")
            if width:
                current_sheet.column_dimensions[get_column_letter(column_index)].width = width
            elif width_default:
                current_sheet.column_dimensions[get_column_letter(column_index)].width = width_default
            # 设置单元格颜色
            color = row_data.get("color")
            if color:
                if type(color) == str:
                    color = PatternFill(fill_type='solid', start_color=color)
                current_cell.fill = color
            font = row_data.get("font")
            if font:
                if type(font) == dict:
                    font = Font(**font)
                current_cell.font = font
            # 插入图片
            image = row_data.get("image", [])
            if image:
                last_img_width = 0
                for img_index, img_path in enumerate(image):
                    last_img_width = self.insert_image(current_sheet, last_img_width, img_path, current_row_index,
                                                       column_index)
            # 合并单元格
            merge = row_data.get('merge')
            if merge:
                current_sheet.merge_cells(merge)

        self.sheet_row_index[sheet_name] += 1

    def add_multi_row_data(self, sheet_name, all_data, is_center=False, add_border=False, width_default=None,
                           height_default=None):
        """添加多行数据"""
        for row_data in all_data:
            self.add_row_data(sheet_name, row_data, is_center=is_center, add_border=add_border,
                              width_default=width_default, height_default=height_default)

    @classmethod
    def merge_same_row_cell(cls, cell_list, key='value', start_index=1, row_index=1):
        """
        合并相同列数据为行数据
        :param cell_list: 行数据
        :param key: 读取key
        :param start_index: 列起始数
        :param row_index: 行数
        :return:
        """
        counter = Counter([d[key] for d in cell_list])
        for cell_item in cell_list:
            if not counter.get(cell_item['value']) or counter.get(cell_item['value']) in [-1, 1] or not cell_item[
                'value']:
                continue
            if cell_item['value'] in counter.keys():
                cell_item.setdefault('merge',
                                     f'{get_column_letter(start_index)}{row_index}:{get_column_letter(start_index - 1 + counter.get(cell_item["value"]))}{row_index}')
                start_index += counter.get(cell_item['value'])
                counter[cell_item['value']] = -1
        return cell_list

    @classmethod
    def merge_same_column_cell(cls, row_data_list, key='value', column_index='A', step=None):
        """
        合并单列相同数据(合并为一行)
        :param row_data_list: 全部数据
        :param key: 读取key
        :param column_index: 需要合并的列 'A'
        :param step: 合并步数   2
        :return:
        """
        column_int_index = column_index_from_string(column_index) - 1
        counter = Counter([d[column_int_index][key] for d in row_data_list])
        merged_row_index = []
        for row_index, row_item in enumerate(row_data_list, start=1):
            for cell_item in row_item:
                if not counter.get(cell_item['value']) or counter.get(cell_item['value']) < 2 or not cell_item[
                    'value']:
                    continue
                if cell_item['value'] in counter.keys():
                    if row_index in merged_row_index:
                        continue
                    if step:
                        end_point = row_index - 1 + step
                        counter[cell_item['value']] -= step
                    else:
                        end_point = row_index - 1 + counter.get(cell_item["value"])
                        counter[cell_item['value']] = -1
                    cell_item.setdefault('merge',
                                         f'{column_index}{row_index}:{column_index}{end_point}')
                    merged_row_index.extend([row_index, end_point])

        return row_data_list

    @classmethod
    def merge_same_column_cell_multi(cls, row_data_list, key='value', column_index='A', step=None):
        """
        合并多列相同数据(合并为一行)
        :param row_data_list:全部数据
        :param key:读取key
        :param column_index:需要合并的列 'A,B'
        :param step:合并步数 {'A': 2, 'B': None}
        :return:
        """
        column_int_index_list = column_index.split(',')
        counter_dict = {}
        merged_row_index_dict = {}
        for column_index in column_int_index_list:
            target_column_index = column_index_from_string(column_index)
            column_int_index = target_column_index - 1
            counter = Counter([d[column_int_index][key] for d in row_data_list])
            counter_dict[target_column_index] = counter
            merged_row_index_dict[target_column_index] = []
        for row_index, row_item in enumerate(row_data_list, start=1):
            for cur_column_index, cell_item in enumerate(row_item, start=1):
                if not counter_dict.get(cur_column_index, {}).get(cell_item['value']) or counter_dict.get(
                        cur_column_index,
                        {}).get(
                    cell_item['value']) < 2 or not cell_item['value']:
                    continue
                if cell_item['value'] in counter_dict.get(cur_column_index, {}).keys():
                    if row_index in merged_row_index_dict.get(cur_column_index, []):
                        continue
                    if step and type(step.get(get_column_letter(cur_column_index))) == int:
                        end_point = row_index - 1 + step.get(get_column_letter(cur_column_index))
                        counter_dict.get(cur_column_index, {})[cell_item['value']] -= step.get(
                            get_column_letter(cur_column_index))
                    else:
                        end_point = row_index - 1 + counter_dict.get(cur_column_index, {}).get(cell_item["value"])
                        counter_dict.get(cur_column_index, {})[cell_item['value']] = -1
                    cell_item.setdefault('merge',
                                         f'{get_column_letter(cur_column_index)}{row_index}:{get_column_letter(cur_column_index)}{end_point}')
                    merged_row_index_dict.get(cur_column_index, []).extend([row_index, end_point])

        return row_data_list

    def save(self):
        self.excel_obj.save(self.excel_name)


class ReadExcelTools:

    def __init__(self, excel_path):
        self.excel_obj = load_workbook(excel_path)

    def load_data(self, sheet_index=0, sheet_name=''):
        """读取数据"""
        cur_sheet = self.excel_obj[sheet_name if sheet_name else self.excel_obj.sheetnames[sheet_index]]
        max_row = cur_sheet.max_row
        max_col = cur_sheet.max_column
        all_data = []
        for row in cur_sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row):
            all_data.append([i.value for i in row])
        return all_data

    def load_data_with_image(self, sheet_index=0, sheet_name=''):
        cur_sheet = self.excel_obj[sheet_name if sheet_name else self.excel_obj.sheetnames[sheet_index]]
        max_row = cur_sheet.max_row
        max_col = cur_sheet.max_column
        all_data = []
        for index, row in enumerate(cur_sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row), start=1):
            all_data.append(
                [i.value if index == 1 else {"text": str(i.value) if i.value else '', "image": []} for i in row])
        for image in cur_sheet._images:
            row, col = image.anchor.to.row + 1, image.anchor.to.col + 1
            all_data[row - 1][col - 1]["image"].append(image)
        return all_data
